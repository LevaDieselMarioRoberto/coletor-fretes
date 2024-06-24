import json 
from openpyxl import Workbook, load_workbook
from datetime import datetime
from posto import Posto
from logger import Logger
from telegram import Telegram
from config import PLANILHA_DIR, PLANILHA_BKP_DIR, JSON_VIBRA_JJ


class Planilha():

    def __init__(self):
        self.arquivo = PLANILHA_DIR
        self.logger = Logger()
    
    def preenche_planilha(self, posto:Posto):
        try:
            planilha = load_workbook(self.arquivo)
        except FileNotFoundError:
            self.logger.log_error("Arquivo não encontrado!")
            planilha = Workbook()
        except PermissionError:
            self.arquivo = PLANILHA_BKP_DIR
            planilha = Workbook()
            self.logger.log_error("Planilha de fretes está aberta!")

            telegram = Telegram()
            telegram.enviar_mensagem("Planilha de fretes indisponível, verifique!")

        sheet = planilha.active

        rows = [5, 6, 7, 8]
        combustiveis = ['etanol', 'gas', 's10', 's500']

        # Preencher células com preços fob
        columns = ['D', 'G', 'J', 'M', 'P', 'S', 'V']
        for col in columns:
            for row, combustivel in zip(rows, combustiveis):
                    sheet[f'{col}{row}'] = posto.precos['fob'][combustivel]

        # Preencher células com preços cif
        columns = ['C', 'F', 'I', 'L', 'O', 'R', 'U']
        volumes = [5, 10, 15, 20, 25, 30, 35]
        for col, vol in zip(columns, volumes):
            for row, combustivel in zip(rows, combustiveis):
                sheet[f'{col}{row}'] = posto.precos[f'{vol}'][combustivel]

        # Preencher células com preços do Janjão
        try:
            combustiveis = ['Etanol', 'Gasolina', 'S10', 'S500']
            rows = [5, 6, 7, 8]
            with open(JSON_VIBRA_JJ, 'r') as janjao_file:
                janjao = json.load(janjao_file)  
                for combustivel, row in zip(combustiveis, rows):
                    sheet[f'X{row}'] = janjao[f'CIF {combustivel}']
                    sheet[f'Y{row}'] = janjao[f'FOB {combustivel}']
            self.logger.log("Preços do Janjão adicionados")
        except Exception as e:
            self.logger.log_error(f"Erro ao usar o arquivo json do Janjão. Erro {e}")

        data_e_hora = datetime.now().strftime("%d/%m-%H:%M")
        sheet['E23'] = data_e_hora

        planilha.save(self.arquivo)
        self.logger.log("Planilha preenchida com sucesso.")
